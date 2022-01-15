import json as json
import os

from openpyxl import Workbook, load_workbook

from settings import FILENAME, WS_NAMES


class Storage:
  def __init__(self) -> None:
    fileExist = os.path.isfile(FILENAME)

    if fileExist:
      self._wb = load_workbook(FILENAME)
    else:
      self._wb = Workbook()

  def checkWS(self):
    for ws in WS_NAMES:
      if ws.value not in self._wb:
        self._wb.create_sheet(ws.value)

  def addItem(self, wsName: WS_NAMES, obj: object):
    valueList = []

    for attribute, value in obj.__dict__.items():
      valueList.append(value)

    self._wb[wsName.value].append(valueList)

  def save(self):
    self._wb.save(FILENAME)
